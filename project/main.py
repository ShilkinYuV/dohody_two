from re import UNICODE
from typing import Pattern
from PyQt5 import QtCore, QtWidgets
from PyQt5 import QtGui
from PyQt5.QtGui import QBrush, QColor, QFont, QIcon, QImage
from openpyxl.workbook.workbook import Workbook
from MainForm import Ui_MainWindow
from PyQt5.QtWidgets import QAbstractItemView, QDesktopWidget, QDialog, QFileDialog, QLabel, QTableWidgetItem, QHeaderView, QMessageBox, QWidget
from AboutWindow import AboutWindows
import sys
import os
from read_excel import Read
import images_rs
import openpyxl
import pyexcel

class MyWindow(QtWidgets.QMainWindow):

    def __init__(self):
        super(MyWindow, self).__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.ui.loadVipiski.clicked.connect(self.open_file)
        self.ui.loadPrilozhenya.clicked.connect(self.open_file)
        self.ui.sverka.clicked.connect(self.run_sverka)
        self.check_one = False
        self.check_two = False
        self.filename = ''
        self.filename_one = ''
        self.filename_two = ''
        self.aboutForm = None
        self.ui.menu.actions()[0].triggered.connect(self.OpenAbout)
        self.ui.menu.actions()[1].triggered.connect(self.sbros)
        self.res_vipiski = []
        self.res_prilozhenia = []
        self.ui.tableWidget.horizontalHeader().setVisible(False)
        self.ui.tableWidget.verticalHeader().setVisible(False)
        self.ui.tableWidget.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.ui.tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.ui.tableWidget.setVisible(False)
        self.ui.sverka.setEnabled(False)
        self.ui.action_Excel.triggered.connect(self.ExportResultsToExcel)
        self.load_one = False
        self.load_two = False
        self.resize(700, 100)
        self.proverka_uspeshna = 0
        self.setWindowIcon(QtGui.QIcon(':roskazna.png'))

    def ExportResultsToExcel(self): 
        wb = Workbook()
        filename, _ = QFileDialog.getSaveFileName(self,'Сохранить файл','', ".xlsx(*.xlsx)")
        self.sheet = wb.active
        row = 1
        col = 1
        for i in range(self.ui.tableWidget.columnCount()):
            for x in range(self.ui.tableWidget.rowCount()):
                try:             
                    teext = str(self.ui.tableWidget.item(x, i).text())
                    color = openpyxl.styles.colors.Color(rgb=self.ui.tableWidget.item(x, i).background().color().name()[1:])
                    self.sheet.cell(row, col).value = teext
                    if color != openpyxl.styles.colors.Color(rgb='000000'):
                        self.sheet.cell(row,col).fill =  openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=color)
                    row += 1
                except AttributeError:
                    row += 1
            row = 1
            col += 1
        self.sheet.column_dimensions['A'].width = 20
        self.sheet.column_dimensions['B'].width = 20
        self.sheet.column_dimensions['C'].width = 20
        self.sheet.column_dimensions['D'].width = 20
        self.sheet.column_dimensions['E'].width = 20
        self.sheet.column_dimensions['F'].width = 20
        self.sheet.column_dimensions['G'].width = 20
        self.sheet.column_dimensions['H'].width = 20
        self.sheet.column_dimensions['I'].width = 30

        self.sheet.merge_cells('A1:D1')
        self.sheet.merge_cells('E1:H1')
                     
        wb.save(filename=filename)

    def sbros(self):
        self.ui.tableWidget.clear()
        self.ui.tableWidget.setRowCount(0)
        self.ui.tableWidget.setColumnCount(0)
        self.ui.loadVipiski.setEnabled(True)
        self.ui.loadPrilozhenya.setEnabled(True)
        self.ui.sverka.setEnabled(False)
        self.ui.tableWidget.setVisible(False)
        self.resize(700, 100)
        self.res_vipiski = []
        self.res_prilozhenia = []

    def open_file(self):

        self.filename = QFileDialog.getOpenFileName(
            None, 'Открыть', os.path.join(os.path.join(
                os.environ['USERPROFILE']), 'Desktop'),
            'Excel Files(*.xls);;Excel Files(*.xlsx)')

        sender = self.sender()
        if str(self.filename) in "('', '')":
            self.ui.statusbar.showMessage('Файл не выбран')
        else:
            if sender.text() == 'ЗАГРУЗИТЬ ВЫПИСКИ':
                self.filename_one = self.filename
                self.check_one = True
                self.check_two = False
                self.new_thread()
                self.ui.loadVipiski.setEnabled(False)

            else:
                self.filename_two = self.filename
                self.check_one = False
                self.check_two = True
                self.new_thread()
                self.ui.loadPrilozhenya.setEnabled(False)
    
    @QtCore.pyqtSlot(list,bool)
    def appendText(self, list, _bool):

        if _bool:
            self.res_vipiski = list
            self.load_one = True
            messagebox = QMessageBox(
                parent=self, text='Выписки загружены')
            messagebox.setWindowTitle('Уведомление')
            messagebox.setStyleSheet(
                'QPushButton{color: rgb(255, 255, 255); font: 75 12pt "Times New Roman";} QLabel{color: rgb(255,255,255); font: 75 12pt "Times New Roman";}')
            messagebox.show()
        else:
            self.res_prilozhenia = list
            self.load_two = True
            messagebox = QMessageBox(
                parent=self, text='Приложения загружены')
            messagebox.setWindowTitle('Уведомление')
            messagebox.setStyleSheet(
                'QPushButton{color: rgb(255, 255, 255); font: 75 12pt "Times New Roman";} QLabel{color: rgb(255,255,255); font: 75 12pt "Times New Roman";}')
            messagebox.show()

        if self.load_one and self.load_two:
            self.ui.sverka.setEnabled(True)


    def new_thread(self):
        if(self.check_one):
            self.my_thread = Read(my_window=self)
            self.my_thread.appendText.connect(self.appendText)
            self.my_thread.start()
        else:
            self.my_thread2 = Read(my_window=self)
            self.my_thread2.appendText.connect(self.appendText)
            self.my_thread2.start()


    def OpenAbout(self):
        if (self.aboutForm != None):
            self.aboutForm.close()
        self.aboutForm = AboutWindows()
        self.aboutForm.show()

    def run_sverka(self):
        self.wb = Workbook()
        self.ui.tableWidget.setVisible(True)
        
            # Максимаьное число строк = выписки + приложения
            # Есть шанс что все они будут не подходить друг другу
        self.ui.tableWidget.setRowCount(len(self.res_vipiski)+len(self.res_prilozhenia))
        
        self.ui.tableWidget.setColumnCount(9)
        font = QFont()
        font.setBold(True)
        font.setPixelSize(14)
        font_t = QFont()
        font_t.setPixelSize(14)        
        newItem = QTableWidgetItem ("ВЫПИСКИ")
        newItem.setFont(font)
        newItem.setTextAlignment(QtCore.Qt.AlignCenter)
        self.ui.tableWidget.setItem(0, 0, newItem)
        newItem = QTableWidgetItem ("ПРИЛОЖЕНИЯ")
        newItem.setFont(font)
        newItem.setTextAlignment(QtCore.Qt.AlignCenter)
        self.ui.tableWidget.setItem(0, 4, newItem)
        newItem = QTableWidgetItem ("РЕЗУЛЬТАТ СВЕРКИ")
        newItem.setFont(font)
        newItem.setTextAlignment(QtCore.Qt.AlignCenter)
        self.ui.tableWidget.setItem(0, 8, newItem)
        
        schet = 1
        for vipiski in self.res_vipiski:
            temp_prilozhenia = []
            for prilozhenia in self.res_prilozhenia:
                if vipiski.licevoy == prilozhenia.licevoy:
                    temp_prilozhenia.append(prilozhenia)
                    prilozhenia.isUsed = True

            if(len(temp_prilozhenia) == 1):
                newItem = QTableWidgetItem ()
                newItem.setText(str(vipiski.licevoy))
                newItem.setFont(font_t)
                self.ui.tableWidget.setItem(schet, 0, newItem)
                newItem = QTableWidgetItem (str(vipiski.summa))
                newItem.setFont(font_t)
                self.ui.tableWidget.setItem(schet, 1, newItem)
                newItem = QTableWidgetItem (str(vipiski.vozvraty))
                newItem.setFont(font_t)                        
                self.ui.tableWidget.setItem(schet, 2, newItem)
                newItem = QTableWidgetItem (str(vipiski.zachety))
                newItem.setFont(font_t)                        
                self.ui.tableWidget.setItem(schet, 3, newItem)

                newItem = QTableWidgetItem (str(temp_prilozhenia[0].licevoy))
                newItem.setFont(font_t)                                 
                self.ui.tableWidget.setItem(schet, 4, newItem)
                newItem = QTableWidgetItem (str(temp_prilozhenia[0].summa))
                newItem.setFont(font_t)                        
                self.ui.tableWidget.setItem(schet, 5, newItem)
                newItem = QTableWidgetItem (str(temp_prilozhenia[0].vozvraty))
                newItem.setFont(font_t)                        
                self.ui.tableWidget.setItem(schet, 6, newItem)
                newItem = QTableWidgetItem (str(temp_prilozhenia[0].zachety))
                newItem.setFont(font_t)                        
                self.ui.tableWidget.setItem(schet, 7, newItem)
                                        
                if(str(vipiski.licevoy) == str(temp_prilozhenia[0].licevoy) and str(vipiski.summa) == str(temp_prilozhenia[0].summa) and str(vipiski.vozvraty) == str(temp_prilozhenia[0].vozvraty) and str(vipiski.zachety) == str(temp_prilozhenia[0].zachety)):
                    
                    if(str(vipiski.summa) == 'Отсутствует' or str(temp_prilozhenia[0].summa) == 'Отсутствует' or str(vipiski.vozvraty) == 'Отсутствует' or str(temp_prilozhenia[0].vozvraty) == 'Отсутствует' or str(vipiski.zachety) == 'Отсутствует' or str(temp_prilozhenia[0].zachety) == 'Отсутствует'):
                        newItem = QTableWidgetItem ("Данные отсутсвуют")
                        newItem.setBackground(QtGui.QColor(220, 220, 170))
                        newItem.setFont(font)
                        newItem.setTextAlignment(QtCore.Qt.AlignCenter)
                        self.ui.tableWidget.setItem(schet, 8, QTableWidgetItem(newItem)) 
                    else:
                        newItem = QTableWidgetItem ("Проверка пройдена")
                        newItem.setBackground(QtGui.QColor(0, 255, 0))
                        newItem.setFont(font)
                        newItem.setTextAlignment(QtCore.Qt.AlignCenter)
                        self.ui.tableWidget.setItem(schet, 8, QTableWidgetItem(newItem))    
                else:
                    newItem = QTableWidgetItem ("Обнаружено несовпадение")
                    newItem.setBackground(QtGui.QColor(255, 0, 0))
                    newItem.setFont(font)
                    newItem.setTextAlignment(QtCore.Qt.AlignCenter)
                    self.ui.tableWidget.setItem(schet, 8, QTableWidgetItem(newItem))
                    self.proverka_uspeshna = self.proverka_uspeshna + 1     
            
            elif(len(temp_prilozhenia) == 0):
                newItem = QTableWidgetItem (str(vipiski.licevoy))
                newItem.setFont(font_t)
                self.ui.tableWidget.setItem(schet, 0, newItem)
                newItem = QTableWidgetItem (str(vipiski.summa))
                newItem.setFont(font_t)
                self.ui.tableWidget.setItem(schet, 1, newItem)
                newItem = QTableWidgetItem (str(vipiski.vozvraty))
                newItem.setFont(font_t)                        
                self.ui.tableWidget.setItem(schet, 2, newItem)
                newItem = QTableWidgetItem (str(vipiski.zachety))
                newItem.setFont(font_t)                        
                self.ui.tableWidget.setItem(schet, 3, newItem)

                newItem = QTableWidgetItem ("Отсутствует")
                newItem.setFont(font_t)                                 
                self.ui.tableWidget.setItem(schet, 4, newItem)
                newItem = QTableWidgetItem ("Отсутствует")
                newItem.setFont(font_t)                        
                self.ui.tableWidget.setItem(schet, 5, newItem)
                newItem = QTableWidgetItem ("Отсутствует")
                newItem.setFont(font_t)                        
                self.ui.tableWidget.setItem(schet, 6, newItem)
                newItem = QTableWidgetItem ("Отсутствует")
                newItem.setFont(font_t)                        
                self.ui.tableWidget.setItem(schet, 7, newItem)

                newItem = QTableWidgetItem ("Отсутствует приложение")
                newItem.setBackground(QtGui.QColor(255, 0, 0))
                newItem.setFont(font)
                newItem.setTextAlignment(QtCore.Qt.AlignCenter)
                self.ui.tableWidget.setItem(schet, 8, QTableWidgetItem(newItem))
                self.proverka_uspeshna = self.proverka_uspeshna + 1 

            else:
                self.flag = False
                for prilozhenie in temp_prilozhenia:
                    temp_prilozhenie = prilozhenie
                    if(str(vipiski.licevoy) == str(prilozhenie.licevoy) and str(vipiski.summa) == str(prilozhenie.summa) and str(vipiski.vozvraty) == str(prilozhenie.vozvraty) and str(vipiski.zachety) == str(prilozhenie.zachety)):
                        self.flag = True
                        newItem = QTableWidgetItem (str(vipiski.licevoy))
                        newItem.setFont(font_t)
                        self.ui.tableWidget.setItem(schet, 0, newItem)
                        newItem = QTableWidgetItem (str(vipiski.summa))
                        newItem.setFont(font_t)
                        self.ui.tableWidget.setItem(schet, 1, newItem)
                        newItem = QTableWidgetItem (str(vipiski.vozvraty))
                        newItem.setFont(font_t)                        
                        self.ui.tableWidget.setItem(schet, 2, newItem)
                        newItem = QTableWidgetItem (str(vipiski.zachety))
                        newItem.setFont(font_t)                        
                        self.ui.tableWidget.setItem(schet, 3, newItem)

                        newItem = QTableWidgetItem (str(prilozhenie.licevoy))
                        newItem.setFont(font_t)                                 
                        self.ui.tableWidget.setItem(schet, 4, newItem)
                        newItem = QTableWidgetItem (str(prilozhenie.summa))
                        newItem.setFont(font_t)                        
                        self.ui.tableWidget.setItem(schet, 5, newItem)
                        newItem = QTableWidgetItem (str(prilozhenie.vozvraty))
                        newItem.setFont(font_t)                        
                        self.ui.tableWidget.setItem(schet, 6, newItem)
                        newItem = QTableWidgetItem (str(prilozhenie.zachety))
                        newItem.setFont(font_t)                        
                        self.ui.tableWidget.setItem(schet, 7, newItem)

                        if(str(vipiski.summa) == 'Отсутствует' or str(prilozhenie.summa) == 'Отсутствует' or str(vipiski.vozvraty) == 'Отсутствует' or str(prilozhenie.vozvraty) == 'Отсутствует' or str(vipiski.zachety) == 'Отсутствует' or str(prilozhenie.zachety) == 'Отсутствует'):
                            newItem = QTableWidgetItem ("Данные отсутсвуют")
                            newItem.setBackground(QtGui.QColor(220, 220, 170))
                            newItem.setFont(font)
                            newItem.setTextAlignment(QtCore.Qt.AlignCenter)
                            self.ui.tableWidget.setItem(schet, 8, QTableWidgetItem(newItem)) 
                        else:
                            newItem = QTableWidgetItem ("Проверка пройдена")
                            newItem.setBackground(QtGui.QColor(220, 220, 170))
                            newItem.setFont(font)
                            newItem.setTextAlignment(QtCore.Qt.AlignCenter)
                            self.ui.tableWidget.setItem(schet, 8, QTableWidgetItem(newItem)) 
                        
                if(self.flag == False):
                    newItem = QTableWidgetItem (str(vipiski.licevoy))
                    newItem.setFont(font_t)
                    self.ui.tableWidget.setItem(schet, 0, newItem)
                    newItem = QTableWidgetItem (str(vipiski.summa))
                    newItem.setFont(font_t)
                    self.ui.tableWidget.setItem(schet, 1, newItem)
                    newItem = QTableWidgetItem (str(vipiski.vozvraty))
                    newItem.setFont(font_t)                        
                    self.ui.tableWidget.setItem(schet, 2, newItem)
                    newItem = QTableWidgetItem (str(vipiski.zachety))
                    newItem.setFont(font_t)                        
                    self.ui.tableWidget.setItem(schet, 3, newItem)

                    newItem = QTableWidgetItem (str(temp_prilozhenie.licevoy))
                    newItem.setFont(font_t)                                 
                    self.ui.tableWidget.setItem(schet, 4, newItem)
                    newItem = QTableWidgetItem (str(temp_prilozhenie.summa))
                    newItem.setFont(font_t)                        
                    self.ui.tableWidget.setItem(schet, 5, newItem)
                    newItem = QTableWidgetItem (str(temp_prilozhenie.vozvraty))
                    newItem.setFont(font_t)                        
                    self.ui.tableWidget.setItem(schet, 6, newItem)
                    newItem = QTableWidgetItem (str(temp_prilozhenie.zachety))
                    newItem.setFont(font_t)                        
                    self.ui.tableWidget.setItem(schet, 7, newItem)
                    
                    newItem = QTableWidgetItem ("Обнаружено несовпадение")
                    newItem.setBackground(QtGui.QColor(255, 0, 0))
                    newItem.setFont(font)
                    newItem.setTextAlignment(QtCore.Qt.AlignCenter)
                    self.ui.tableWidget.setItem(schet, 8, QTableWidgetItem(newItem))
                    self.proverka_uspeshna = self.proverka_uspeshna + 1   


            schet = schet + 1
        
        for prilozh in self.res_prilozhenia:
            if(prilozh.isUsed != True):
                print(schet)
                newItem = QTableWidgetItem ("Отсутствует")
                newItem.setFont(font_t)
                self.ui.tableWidget.setItem(schet, 0, QTableWidgetItem(newItem))
                newItem = QTableWidgetItem ("Отсутствует")
                newItem.setFont(font_t)
                self.ui.tableWidget.setItem(schet, 1, QTableWidgetItem(newItem))
                newItem = QTableWidgetItem ("Отсутствует")
                newItem.setFont(font_t)                        
                self.ui.tableWidget.setItem(schet, 2, QTableWidgetItem(newItem))
                newItem = QTableWidgetItem ("Отсутствует")
                newItem.setFont(font_t)                        
                self.ui.tableWidget.setItem(schet, 3, QTableWidgetItem(newItem))
                
                newItem = QTableWidgetItem (str(prilozh.licevoy))
                newItem.setFont(font_t)                                 
                self.ui.tableWidget.setItem(schet, 4, QTableWidgetItem(newItem))
                newItem = QTableWidgetItem (str(prilozh.summa))
                newItem.setFont(font_t)                        
                self.ui.tableWidget.setItem(schet, 5, QTableWidgetItem(newItem))
                newItem = QTableWidgetItem (str(prilozh.vozvraty))
                newItem.setFont(font_t)                        
                self.ui.tableWidget.setItem(schet, 6, QTableWidgetItem(newItem))
                newItem = QTableWidgetItem (str(prilozh.zachety))
                newItem.setFont(font_t)                        
                self.ui.tableWidget.setItem(schet, 7, QTableWidgetItem(newItem))

                newItem = QTableWidgetItem ("Отсутсвует выписка")
                newItem.setBackground(QtGui.QColor(255, 0, 0))
                newItem.setFont(font)
                newItem.setTextAlignment(QtCore.Qt.AlignCenter)
                self.ui.tableWidget.setItem(schet, 8, QTableWidgetItem(newItem))
                self.proverka_uspeshna = self.proverka_uspeshna + 1 

                schet = schet + 1
                
        if(self.proverka_uspeshna == 0):
            messagebox = QMessageBox(
                parent=self, text='ВСЕ ДАННЫЕ СОВПАДАЮТ!!!')
            messagebox.setWindowTitle('Результат проверки')
            messagebox.setStyleSheet(
                'QPushButton{color: rgb(255, 255, 255); font: 75 12pt "Times New Roman";} QLabel{color: rgb(255,255,255); font: 75 12pt "Times New Roman";}')
            messagebox.show()
        else:
            messagebox = QMessageBox(
                parent=self, text='ОБНАРУЖЕНЫ НЕСОВПАДЕНИЯ!!!')
            messagebox.setWindowTitle('Результат проверки')
            messagebox.setStyleSheet(
                'QPushButton{color: rgb(255, 255, 255); font: 75 12pt "Times New Roman";} QLabel{color: rgb(255,255,255); font: 75 12pt "Times New Roman";}')
            messagebox.show()
        self.ui.tableWidget.setSpan(0,0,1,4)
        self.ui.tableWidget.setSpan(0,4,1,4)
        self.resize(1400, 600)
        qtRectangle = self.frameGeometry()
        centerPoint = QDesktopWidget().availableGeometry().center()
        qtRectangle.moveCenter(centerPoint)
        self.move(qtRectangle.topLeft())
        self.ui.sverka.setEnabled(False)
        

app = QtWidgets.QApplication([])
application = MyWindow()
application.setWindowTitle("Сверщик выписок")
application.show()

sys.exit(app.exec())
